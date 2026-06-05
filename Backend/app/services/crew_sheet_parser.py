"""解析演职员表 Excel（职位 | 人员1 | 人员2 …）。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.exceptions import BizError

# 与前台 translations.ts 中职位/字段名对齐
_ROLE_TRANSLATIONS: dict[str, dict[str, str]] = {
    "导演": {"CN": "导演", "EN": "Director", "JP": "監督"},
    "编剧": {"CN": "编剧", "EN": "Screenwriter", "JP": "脚本"},
    "摄影指导": {"CN": "摄影指导", "EN": "Cinematographer", "JP": "撮影監督"},
    "后期剪接": {"CN": "后期剪接", "EN": "Editor", "JP": "編集"},
    "调色": {"CN": "调色", "EN": "Colorist", "JP": "カラリスト"},
    "剪辑": {"CN": "剪辑", "EN": "Editor", "JP": "編集"},
    "制片人": {"CN": "制片人", "EN": "Producer", "JP": "プロデューサー"},
    "演员": {"CN": "演员", "EN": "Cast", "JP": "キャスト"},
    "音效": {"CN": "音效", "EN": "Sound", "JP": "音響"},
    "美术指导": {"CN": "美术指导", "EN": "Art Director", "JP": "美術監督"},
    "音乐": {"CN": "音乐", "EN": "Music", "JP": "音楽"},
    "director": {"CN": "导演", "EN": "Director", "JP": "監督"},
    "producer": {"CN": "制片人", "EN": "Producer", "JP": "プロデューサー"},
    "cast": {"CN": "演员", "EN": "Cast", "JP": "キャスト"},
    "screenwriter": {"CN": "编剧", "EN": "Screenwriter", "JP": "脚本"},
    "cinematographer": {"CN": "摄影指导", "EN": "Cinematographer", "JP": "撮影監督"},
    "sound": {"CN": "音效", "EN": "Sound", "JP": "音響"},
    "artDirector": {"CN": "美术指导", "EN": "Art Director", "JP": "美術監督"},
    "editor": {"CN": "剪辑", "EN": "Editor", "JP": "編集"},
    "music": {"CN": "音乐", "EN": "Music", "JP": "音楽"},
}

_HEADER_ROLE_MARKERS = frozenset(
    {
        "职位",
        "職位",
        "role",
        "position",
        "position / role",
        "职务",
    }
)
_HEADER_NAME_MARKERS = frozenset({"人员", "人員", "name", "names", "cast", "staff"})
_SKIP_NAME_VALUES = frozenset({"...", "…", "-", "—", "n/a", "na", "none", ""})
_MAX_FILE_BYTES = 2 * 1024 * 1024
_ALLOWED_SUFFIXES = (".xlsx", ".xlsm")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def localize_role(raw_role: str) -> dict[str, str]:
    role = _normalize_cell(raw_role)
    if not role:
        return {"CN": "", "EN": "", "JP": ""}
    if role in _ROLE_TRANSLATIONS:
        entry = _ROLE_TRANSLATIONS[role]
        return {"CN": entry["CN"], "EN": entry["EN"], "JP": entry.get("JP", entry["EN"])}
    lowered = role.lower()
    if lowered in _ROLE_TRANSLATIONS:
        entry = _ROLE_TRANSLATIONS[lowered]
        return {"CN": entry["CN"], "EN": entry["EN"], "JP": entry.get("JP", entry["EN"])}
    return {"CN": role, "EN": role, "JP": role}


def _is_header_row(row: tuple[Any, ...]) -> bool:
    if not row:
        return False
    first = _normalize_cell(row[0]).lower()
    if first in _HEADER_ROLE_MARKERS:
        return True
    joined = " ".join(_normalize_cell(cell).lower() for cell in row[:3])
    return any(marker in joined for marker in _HEADER_ROLE_MARKERS) and any(
        marker in joined for marker in _HEADER_NAME_MARKERS
    )


def parse_crew_sheet_bytes(content: bytes, *, filename: str = "") -> list[dict[str, Any]]:
    if not content:
        raise BizError("文件为空")
    if len(content) > _MAX_FILE_BYTES:
        raise BizError("演职员表文件不能超过 2MB")

    suffix = filename.lower() if filename else ""
    if suffix and not any(suffix.endswith(ext) for ext in _ALLOWED_SUFFIXES):
        raise BizError("仅支持 .xlsx / .xlsm 格式的 Excel 文件")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BizError(f"无法读取 Excel 文件：{exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise BizError("表格中没有数据")

    start_index = 1 if _is_header_row(rows[0]) else 0
    crew_credits: list[dict[str, Any]] = []

    for row in rows[start_index:]:
        if not row:
            continue
        role_text = _normalize_cell(row[0])
        if not role_text or role_text.lower() in _HEADER_ROLE_MARKERS:
            continue

        names: list[str] = []
        for cell in row[1:]:
            name = _normalize_cell(cell)
            if not name or name.lower() in _SKIP_NAME_VALUES:
                continue
            names.append(name)

        if not names:
            continue

        role = localize_role(role_text)
        if not role["CN"]:
            continue

        crew_credits.append({"role": role, "names": names})

    if not crew_credits:
        raise BizError("未解析到有效的职位与人员，请检查表格格式（A 列为职位，B 列起为人员）")

    return crew_credits
